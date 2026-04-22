import pytest
from openpyxl import Workbook
from src.sharepoint.import_inventory import (
    HEADER_MAP,
    load_headers,
    validate_headers,
)


def _sheet_with_headers(headers):
    wb = Workbook()
    ws = wb.active
    for col, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=h)
    return ws


def test_load_headers_maps_names_to_column_indexes():
    ws = _sheet_with_headers(["Device name", "Serial number", "Make"])
    result = load_headers(ws)
    assert result == {"device name": 1, "serial number": 2, "make": 3}


def test_load_headers_is_case_insensitive_and_trims_whitespace():
    ws = _sheet_with_headers(["  DEVICE NAME  ", " Make "])
    result = load_headers(ws)
    assert result["device name"] == 1
    assert result["make"] == 2


def test_validate_headers_passes_when_all_required_present():
    headers = {k.lower(): i for i, k in enumerate(HEADER_MAP.keys(), start=1)}
    validate_headers(headers)  # must not raise


def test_validate_headers_raises_listing_missing_and_found():
    headers = {"device name": 1, "make": 2}
    with pytest.raises(ValueError) as exc:
        validate_headers(headers)
    msg = str(exc.value)
    assert "Missing required headers" in msg
    assert "Serial number" in msg
    assert "Found headers" in msg
    assert "Make" in msg


from src.sharepoint.import_inventory import (
    BYTES_PER_GB,
    transform_row,
    iter_rows,
)


def _full_headers():
    return {
        "device name": 1,
        "serial number": 2,
        "make": 3,
        "model": 4,
        "disk 1 size": 5,
        "total physical memory": 6,
        "cpu name": 7,
    }


def test_transform_row_strips_text_and_converts_gb():
    row = ("  EW22-01322 ", "ABC123", "Dell", "LATITUDE 5530",
           256 * BYTES_PER_GB, 16 * BYTES_PER_GB, "Intel Core i5-1235U")
    result = transform_row(row, _full_headers())
    assert result == {
        "DeviceName": "EW22-01322",
        "SerialNumber": "ABC123",
        "Make": "Dell",
        "Model": "LATITUDE 5530",
        "DiskSize": 256,
        "RAM": 16,
        "CPU": "Intel Core i5-1235U",
    }


def test_transform_row_rounds_nearest_gb():
    # 15.9 GB of bytes -> 16 after rounding
    almost_16 = int(15.9 * BYTES_PER_GB)
    row = ("EW22-0", "S", "Dell", "L", almost_16, almost_16, "i5")
    result = transform_row(row, _full_headers())
    assert result["DiskSize"] == 16
    assert result["RAM"] == 16


def test_transform_row_returns_none_when_device_name_blank():
    row = ("   ", "S", "Dell", "L", 100, 100, "i5")
    assert transform_row(row, _full_headers()) is None


def test_transform_row_handles_non_numeric_spec_by_emitting_zero():
    row = ("EW22-0", "S", "Dell", "L", "N/A", None, "i5")
    result = transform_row(row, _full_headers())
    assert result["DiskSize"] == 0
    assert result["RAM"] == 0


def test_iter_rows_yields_dicts_and_respects_limit(tmp_path):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(list(_full_headers().keys()))  # headers as-is
    for i in range(5):
        ws.append([f"EW22-0000{i}", f"SN{i}", "Dell", "LATITUDE 5530",
                   256 * BYTES_PER_GB, 16 * BYTES_PER_GB, "i5"])
    xlsx = tmp_path / "sample.xlsx"
    wb.save(xlsx)

    rows = list(iter_rows(str(xlsx), limit=3))
    assert len(rows) == 3
    assert rows[0]["DeviceName"] == "EW22-00000"
    assert rows[2]["DeviceName"] == "EW22-00002"


def test_iter_rows_skips_blank_device_name_rows(tmp_path):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(list(_full_headers().keys()))
    ws.append(["EW22-00001", "SN1", "Dell", "L", 256 * BYTES_PER_GB, 16 * BYTES_PER_GB, "i5"])
    ws.append(["", "SN2", "Dell", "L", 256 * BYTES_PER_GB, 16 * BYTES_PER_GB, "i5"])
    ws.append(["EW22-00003", "SN3", "Dell", "L", 256 * BYTES_PER_GB, 16 * BYTES_PER_GB, "i5"])
    xlsx = tmp_path / "sample.xlsx"
    wb.save(xlsx)

    rows = list(iter_rows(str(xlsx)))
    names = [r["DeviceName"] for r in rows]
    assert names == ["EW22-00001", "EW22-00003"]
