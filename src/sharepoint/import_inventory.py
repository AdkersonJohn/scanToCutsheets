"""
import_inventory.py — one-shot loader for RefreshAssetInventory SharePoint list.

Reads an Excel export and pushes rows to SharePoint via REST $batch using
user-delegated credentials (device code flow). No admin consent required.

Run via: npm run inventory
"""

from __future__ import annotations

import json
import sys
import time
import uuid
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterator

import msal
import requests
from openpyxl import load_workbook


# --- Constants ---

TENANT_URL = "https://encoretch.sharepoint.com"
SITE_PATH = "/sites/CCHMCRefreshSupport"
SITE_URL = f"{TENANT_URL}{SITE_PATH}"
LIST_NAME = "RefreshAssetInventory"

# Pre-consented Microsoft public apps (device code flow, no admin needed).
CLIENT_IDS = [
    "14d82eec-204b-4c2f-b7e8-296a70dab67e",  # Microsoft Graph CLI Tools
    "31359c7f-bd7e-475c-86db-fdb8c937548e",  # PnP Management Shell
]
SCOPES = [f"{TENANT_URL}/.default"]

BATCH_SIZE = 100
BYTES_PER_GB = 1073741824

# Source Excel header -> (destination SP column, transform kind).
# Transforms: "text" = str(v).strip(); "gb" = round(float(v) / BYTES_PER_GB).
HEADER_MAP: dict[str, tuple[str, str]] = {
    "Device name":            ("DeviceName",   "text"),
    "Serial number":          ("SerialNumber", "text"),
    "Make":                   ("Make",         "text"),
    "Model":                  ("Model",        "text"),
    "Disk 1 size":            ("DiskSize",     "gb"),
    "Total physical memory":  ("RAM",          "gb"),
    "CPU name":               ("CPU",          "text"),
}

CACHE_PATH = Path(__file__).parent / ".token_cache.json"
PROGRESS_PATH = Path(__file__).parent / ".import_progress.json"


# --- Excel ---

def load_headers(sheet) -> dict[str, int]:
    """Read row 1 as headers, return {normalized_header: 1-based column index}."""
    result: dict[str, int] = {}
    for cell in sheet[1]:
        if cell.value is None:
            continue
        key = str(cell.value).strip().lower()
        if key:
            result[key] = cell.column
    return result


def validate_headers(headers: dict[str, int]) -> None:
    """Raise ValueError if any required header from HEADER_MAP is missing."""
    missing = [name for name in HEADER_MAP if name.lower() not in headers]
    if missing:
        # Map normalized headers back to their original HEADER_MAP names for display
        original_map = {k.lower(): k for k in HEADER_MAP}
        found = sorted([original_map.get(h, h) for h in headers.keys()])
        raise ValueError(
            f"Missing required headers: {missing}. Found headers: {found}"
        )


def transform_row(row_values: tuple, header_index: dict[str, int]) -> dict | None:
    """
    Transform a raw Excel row tuple into a SharePoint item dict.
    Returns None for rows that should be skipped (blank DeviceName).
    """
    out: dict = {}
    for src_header, (dest_col, kind) in HEADER_MAP.items():
        col_idx = header_index.get(src_header.lower())
        if col_idx is None:
            # validate_headers should have caught this; defensive fallback.
            out[dest_col] = "" if kind == "text" else 0
            continue
        raw = row_values[col_idx - 1]
        if kind == "text":
            out[dest_col] = "" if raw is None else str(raw).strip()
        elif kind == "gb":
            try:
                out[dest_col] = round(float(raw) / BYTES_PER_GB)
            except (TypeError, ValueError):
                out[dest_col] = 0

    if not out.get("DeviceName"):
        return None
    return out


def iter_rows(xlsx_path: str, limit: int | None = None) -> Iterator[dict]:
    """
    Stream transformed rows from the workbook. Validates headers on open.
    Skips rows where DeviceName is blank. Respects `limit` if given (counts
    emitted rows, not skipped ones).
    """
    wb = load_workbook(filename=xlsx_path, read_only=True, data_only=True)
    try:
        sheet = wb.active
        headers = load_headers(sheet)
        validate_headers(headers)

        emitted = 0
        for idx, row in enumerate(sheet.iter_rows(values_only=True)):
            if idx == 0:
                continue  # header row
            transformed = transform_row(row, headers)
            if transformed is None:
                continue
            yield transformed
            emitted += 1
            if limit is not None and emitted >= limit:
                return
    finally:
        wb.close()


# --- Batch Operations ---


def _list_item_type(list_name: str) -> str:
    # SharePoint list item content type: SP.Data.<ListName>ListItem
    return f"SP.Data.{list_name}ListItem"


def build_batch_body(items: list[dict], list_name: str) -> tuple[str, str]:
    """
    Build a multipart/mixed $batch body containing a single changeset with one
    POST per item. Returns (body, outer_boundary).
    """
    batch_boundary = f"batch_{uuid.uuid4()}"
    changeset_boundary = f"changeset_{uuid.uuid4()}"
    list_url = f"{SITE_URL}/_api/web/lists/getbytitle('{list_name}')/items"
    item_type = _list_item_type(list_name)

    lines: list[str] = []
    lines.append(f"--{batch_boundary}")
    lines.append(f"Content-Type: multipart/mixed; boundary={changeset_boundary}")
    lines.append("Content-Length: 1")
    lines.append("Content-Transfer-Encoding: binary")
    lines.append("")

    for idx, item in enumerate(items, start=1):
        payload = {"__metadata": {"type": item_type}, **item}
        body_json = json.dumps(payload, separators=(",", ":"))
        lines.append(f"--{changeset_boundary}")
        lines.append("Content-Type: application/http")
        lines.append("Content-Transfer-Encoding: binary")
        lines.append(f"Content-ID: {idx}")
        lines.append("")
        lines.append(f"POST {list_url} HTTP/1.1")
        lines.append("Accept: application/json;odata=verbose")
        lines.append("Content-Type: application/json;odata=verbose")
        lines.append("")
        lines.append(body_json)
        lines.append("")

    lines.append(f"--{changeset_boundary}--")
    lines.append(f"--{batch_boundary}--")
    lines.append("")
    return "\r\n".join(lines), batch_boundary


def build_delete_batch_body(item_ids: list[int], list_name: str) -> tuple[str, str]:
    """Build a $batch body that deletes each given item id."""
    batch_boundary = f"batch_{uuid.uuid4()}"
    changeset_boundary = f"changeset_{uuid.uuid4()}"
    list_url = f"{SITE_URL}/_api/web/lists/getbytitle('{list_name}')/items"

    lines: list[str] = []
    lines.append(f"--{batch_boundary}")
    lines.append(f"Content-Type: multipart/mixed; boundary={changeset_boundary}")
    lines.append("Content-Length: 1")
    lines.append("Content-Transfer-Encoding: binary")
    lines.append("")

    for idx, item_id in enumerate(item_ids, start=1):
        lines.append(f"--{changeset_boundary}")
        lines.append("Content-Type: application/http")
        lines.append("Content-Transfer-Encoding: binary")
        lines.append(f"Content-ID: {idx}")
        lines.append("")
        lines.append(f"DELETE {list_url}({item_id}) HTTP/1.1")
        lines.append("Accept: application/json;odata=verbose")
        lines.append("If-Match: *")
        lines.append("")

    lines.append(f"--{changeset_boundary}--")
    lines.append(f"--{batch_boundary}--")
    lines.append("")
    return "\r\n".join(lines), batch_boundary


# --- HTTP & Retry ---


class RetryExhaustedError(RuntimeError):
    """Raised when a $batch call fails all retries."""


def sp_headers(token: str, extra: dict | None = None) -> dict:
    headers = {
        "Authorization": f"Bearer {token}",
        "Accept": "application/json;odata=verbose",
    }
    if extra:
        headers.update(extra)
    return headers


def list_item_count(token: str) -> int:
    url = f"{SITE_URL}/_api/web/lists/getbytitle('{LIST_NAME}')/ItemCount"
    resp = requests.get(url, headers=sp_headers(token), timeout=60)
    resp.raise_for_status()
    return int(resp.json()["d"]["ItemCount"])


def list_all_item_ids(token: str) -> list[int]:
    """Return every item Id in the list, following OData paging."""
    base = f"{SITE_URL}/_api/web/lists/getbytitle('{LIST_NAME}')/items"
    params = {"$select": "Id", "$top": "5000"}
    ids: list[int] = []
    url = base
    while url:
        resp = requests.get(url, headers=sp_headers(token),
                            params=params if url == base else None,
                            timeout=120)
        resp.raise_for_status()
        data = resp.json()["d"]
        ids.extend(int(r["Id"]) for r in data.get("results", []))
        url = data.get("__next")
    return ids


def post_batch(body: str, boundary: str, token: str,
               max_retries: int = 5, base_delay: float = 2.0) -> None:
    """
    POST a multipart/mixed $batch body to SharePoint with retry/backoff.
    Retries 429, 5xx, and network errors (ConnectionError, Timeout);
    raises RetryExhaustedError on any other failure or after max_retries exhausted.
    """
    url = f"{SITE_URL}/_api/$batch"
    headers = sp_headers(token, {
        "Content-Type": f"multipart/mixed; boundary={boundary}",
    })

    attempt = 0
    while True:
        try:
            resp = requests.post(url, headers=headers, data=body.encode("utf-8"), timeout=120)
        except (requests.exceptions.ConnectionError, requests.exceptions.Timeout) as exc:
            if attempt >= max_retries:
                raise RetryExhaustedError(
                    f"Batch POST failed due to network error: {exc}"
                ) from exc
            delay = base_delay * (2 ** attempt)
            time.sleep(delay)
            attempt += 1
            continue

        if 200 <= resp.status_code < 300:
            return
        should_retry = resp.status_code == 429 or resp.status_code >= 500
        if not should_retry or attempt >= max_retries:
            raise RetryExhaustedError(
                f"Batch POST failed: HTTP {resp.status_code}. Body: {resp.text[:500]}"
            )
        retry_after = resp.headers.get("Retry-After")
        if retry_after is not None:
            try:
                delay = float(retry_after)
            except ValueError:
                delay = base_delay * (2 ** attempt)
        else:
            delay = base_delay * (2 ** attempt)
        time.sleep(delay)
        attempt += 1


# --- Progress ---


def save_progress(phase: str, last_batch: int, total: int) -> None:
    """Save progress state to disk, preserving started_at if already set."""
    existing = load_progress() or {}
    started_at = existing.get("started_at") or datetime.now(timezone.utc).isoformat()
    data = {
        "phase": phase,
        "last_batch": last_batch,
        "total": total,
        "started_at": started_at,
    }
    PROGRESS_PATH.write_text(json.dumps(data, indent=2))


def load_progress() -> dict | None:
    """Load progress state from disk, returning None if file missing or invalid."""
    if not PROGRESS_PATH.exists():
        return None
    try:
        return json.loads(PROGRESS_PATH.read_text())
    except (OSError, json.JSONDecodeError):
        return None


def clear_progress() -> None:
    """Delete progress file, silently succeeding if file does not exist."""
    try:
        PROGRESS_PATH.unlink()
    except FileNotFoundError:
        pass


# --- Auth ---

AUTH_AUTHORITY = "https://login.microsoftonline.com/common"


def _build_msal_app(client_id: str):
    cache = msal.SerializableTokenCache()
    if CACHE_PATH.exists():
        try:
            cache.deserialize(CACHE_PATH.read_text())
        except Exception:
            pass
    app = msal.PublicClientApplication(
        client_id,
        authority=AUTH_AUTHORITY,
        token_cache=cache,
    )
    app._token_cache_ref = cache  # keep a ref for save-on-exit
    return app


def _save_cache(app) -> None:
    cache = getattr(app, "_token_cache_ref", None)
    if cache is not None and cache.has_state_changed:
        CACHE_PATH.write_text(cache.serialize())


def _try_client(client_id: str) -> str | None:
    app = _build_msal_app(client_id)
    # Try silent token acquisition first.
    for acct in app.get_accounts():
        result = app.acquire_token_silent(SCOPES, account=acct)
        if result and "access_token" in result:
            _save_cache(app)
            return result["access_token"]

    flow = app.initiate_device_flow(scopes=SCOPES)
    if "user_code" not in flow:
        return None
    print(flow.get("message", "Visit https://microsoft.com/devicelogin"))
    sys.stdout.flush()
    result = app.acquire_token_by_device_flow(flow)
    if "access_token" in result:
        _save_cache(app)
        return result["access_token"]
    # Print the error so the user can distinguish "I cancelled" from "blocked".
    print(f"Auth failed with client {client_id}: "
          f"{result.get('error')} — {result.get('error_description')}",
          file=sys.stderr)
    return None


def acquire_token() -> str:
    """
    Try each client ID in CLIENT_IDS in order. Return the first access token
    obtained. Raises RuntimeError if all fail.
    """
    for client_id in CLIENT_IDS:
        token = _try_client(client_id)
        if token:
            return token
    raise RuntimeError(
        "Could not acquire a SharePoint token with any public client ID. "
        "Ask your site owner to grant consent to "
        f"one of: {CLIENT_IDS}"
    )


# --- Orchestration ---

import argparse


def _chunks(seq: list, n: int):
    for i in range(0, len(seq), n):
        yield seq[i:i + n]


def clear_list(token: str) -> int:
    ids = list_all_item_ids(token)
    if not ids:
        return 0
    total = len(ids)
    print(f"Deleting {total} existing items...")
    for i, chunk in enumerate(_chunks(ids, BATCH_SIZE), start=1):
        body, boundary = build_delete_batch_body(chunk, LIST_NAME)
        post_batch(body, boundary, token)
        done = min(i * BATCH_SIZE, total)
        print(f"  deleted {done}/{total}")
    return total


def import_items(token: str, rows, resume_from: int = 0) -> None:
    buffer: list[dict] = []
    batch_idx = 0
    sent = 0

    for row in rows:
        buffer.append(row)
        if len(buffer) >= BATCH_SIZE:
            batch_idx += 1
            if batch_idx <= resume_from:
                buffer.clear()
                continue
            body, boundary = build_batch_body(buffer, LIST_NAME)
            post_batch(body, boundary, token)
            sent += len(buffer)
            save_progress("insert", batch_idx, total=-1)
            if batch_idx % 10 == 0:
                print(f"  inserted batch {batch_idx} ({sent} items)")
            buffer.clear()

    if buffer:
        batch_idx += 1
        if batch_idx > resume_from:
            body, boundary = build_batch_body(buffer, LIST_NAME)
            post_batch(body, boundary, token)
            sent += len(buffer)
            save_progress("insert", batch_idx, total=-1)

    print(f"Inserted {sent} items across {batch_idx} batches.")


def _confirm(prompt: str) -> bool:
    try:
        ans = input(prompt).strip().lower()
    except EOFError:
        return False
    return ans == "yes"


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Load Refresh Asset Data.xlsx into SharePoint RefreshAssetInventory."
    )
    parser.add_argument("--file", required=True, help="Path to Refresh Asset Data.xlsx")
    parser.add_argument("--clear-first", action="store_true",
                        help="Delete all existing items before importing")
    parser.add_argument("--resume", action="store_true",
                        help="Resume an interrupted insert phase")
    parser.add_argument("--dry-run", action="store_true",
                        help="Print first 5 transformed rows, post nothing")
    parser.add_argument("--limit", type=int, default=None,
                        help="Only process the first N rows")
    args = parser.parse_args()

    xlsx = Path(args.file)
    if not xlsx.exists():
        print(f"Excel file not found: {xlsx}", file=sys.stderr)
        return 2

    if args.dry_run:
        print(f"Reading {xlsx} (dry run)...")
        sample = []
        for row in iter_rows(str(xlsx), limit=args.limit or 5):
            sample.append(row)
            if len(sample) >= 5:
                break
        for r in sample:
            print(json.dumps(r, indent=2))
        print(f"Dry run OK. {len(sample)} rows shown.")
        return 0

    print("Authenticating...")
    token = acquire_token()
    print("Authenticated.")

    count = list_item_count(token)
    print(f"List currently has {count} items.")

    resume_from = 0
    if args.resume:
        state = load_progress()
        if not state or state.get("phase") != "insert":
            print("No insert phase to resume.", file=sys.stderr)
            return 2
        resume_from = int(state.get("last_batch", 0))
        print(f"Resuming from batch {resume_from}.")
    elif args.clear_first:
        if count > 0:
            if not _confirm(
                f"This will delete all {count} existing items in {LIST_NAME}. "
                "Type 'yes' to continue: "):
                print("Aborted.")
                return 1
            clear_list(token)
        clear_progress()
    else:
        if count > 0:
            print(f"List has {count} items. Use --clear-first to replace, "
                  "or empty the list manually first.", file=sys.stderr)
            return 2

    print(f"Reading {xlsx}...")
    rows = iter_rows(str(xlsx), limit=args.limit)
    import_items(token, rows, resume_from=resume_from)
    clear_progress()
    print("Done.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
